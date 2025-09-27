from datetime import datetime
from pathlib import Path
from typing import Optional, TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from app.config import settings
from app.crud import get_documents_with_grouped_plans, get_documents_count
from app.db import get_db
from app.utils.base import format_string_list, get_localized_months_list
from app.utils.console import print_success, print_error, print_warning, console
from app.utils.files import get_export_file_path, cleanup_existing_files, get_unique_filename

if TYPE_CHECKING:
	from app.models import Document


def export_documents_to_file(
		year: int,
		output_dir: Path,
		rows_per_file: int = settings.MAX_DOCUMENTS_PER_EXPORT_FILE,
		force_update: bool = settings.REWRITE_FILE_ON_CONFLICT,
		offset: Optional[int] = None,
		limit: Optional[int] = None,
		title: str = "Экспорт документов"
) -> tuple[list[Path], int]:
	"""
	Экспорт документов в XLSX с разбиением на части.

	Args:
		year: Год для фильтрации
		output_dir: Папка для сохранения
		rows_per_file: Сколько документов в одном файле (0 = все в один файл)
		force_update: Принудительная перезапись файлов
		offset: Смещение для начала выборки
		limit: Максимальное количество документов для экспорта
		title: Заголовок для вывода в консоль

	Returns:
		(export_paths, total_exported)
	"""

	console.print("=" * len(title), style="blue")
	console.print(title.upper(), style="bold blue")
	console.print("=" * len(title) + '\n', style="blue")

	output_dir.mkdir(exist_ok=True)
	export_paths: list[Path] = []
	total_exported = 0
	current_offset = offset or 0
	part_num = 1

	with next(get_db()) as db:
		# Если указан лимит или диапазон, делаем COUNT для валидации
		if limit is not None or (offset is not None and offset > 0):
			total = get_documents_count(db, year=year)

			if not total:
				print_warning(f"Нет сохранённых документов за {year} год!")
				return [], 0

			if current_offset >= total:
				print_error(f"Смещение {current_offset} превышает количество документов ({total})")
				return [], 0

			if limit is None:
				limit = total - current_offset

		# Если rows_per_file == 0 → всё в один файл
		effective_rows_per_file = rows_per_file or (limit if limit else 0)

		while True:
			# Считаем размер текущего батча
			batch_size = (
				min(effective_rows_per_file, limit - total_exported)
				if limit
				else effective_rows_per_file
			)

			batch_docs = get_documents_with_grouped_plans(
				db,
				year=year,
				skip=current_offset,
				limit=batch_size if batch_size > 0 else None,
			)

			if not batch_docs:
				break

			# Разбиваем на чанки по rows_per_file
			chunks = (
				[batch_docs[i:i + effective_rows_per_file] for i in range(0, len(batch_docs), effective_rows_per_file)]
				if effective_rows_per_file and len(batch_docs) > effective_rows_per_file
				else [batch_docs]
			)

			for i, chunk in enumerate(chunks, 1):
				if not chunk:
					continue

				postfix = (
					f"-part{part_num}"
					if (rows_per_file and (limit is None or limit > rows_per_file)) or len(chunks) > 1
					else ""
				)

				export_path = export_plans_to_xls(
					list(chunk),
					year,
					output_dir,
					postfix,
					force_update,
				)

				if not export_path:
					print_error("Ошибка сохранения файла")
					break

				console.print(
					f"{part_num}: [cyan bold]{export_path}[/cyan bold] "
					f"(записей: {len(chunk)})"
				)

				export_paths.append(export_path)
				total_exported += len(chunk)
				part_num += 1

				if limit and total_exported >= limit:
					break

			current_offset += len(batch_docs)

			if limit and total_exported >= limit:
				break

	# Вывод итоговой информации
	if export_paths:
		if len(export_paths) > 1:
			console.print(f"Создано файлов: [cyan bold]{len(export_paths)}[/cyan bold]")

		console.print("\n" + "=" * 80, style="dim")
		print_success("Экспорт успешно завершен.")
		console.print(f"\nВсего экспортировано документов: [cyan bold]{total_exported}[/cyan bold]")
		console.print("📂 Ссылки на файлы XLSX:", style="dim")
		for path in export_paths:
			console.print(f"   - [blue][link={path}]{path}[/link][/blue]")

		console.print("=" * 80, style="dim")

	return export_paths, total_exported


def export_plans_to_xls(
		documents: list[tuple['Document', dict[str, list[float | None]]]],
		year: int,
		export_dir: Optional[Path] = None,
		postfix: str = "",
		force_overwrite: bool = False
) -> Path:
	"""
	Экспортирует список документов в XLS файл с детализацией по месяцам.
	"""

	# Создаем рабочую книгу и лист
	wb = Workbook()
	ws = wb.active
	ws.title = f"{year}"

	# Убираем сетку листа
	ws.sheet_view.showGridLines = False

	# Добавляем заголовок
	ws['A1'] = f"Сводные данные плановых закупок по контрагентам за {year} год"
	ws['A1'].font = Font(bold=True, size=18)

	# Добавляем дату создания
	creation_date = datetime.now().strftime("%d.%m.%Y %H:%M")
	ws['A2'] = f"Дата формирования: {creation_date}"
	ws['A2'].font = Font(italic=True)

	# Создаем шапку таблицы (начинаем с 4 строки)
	headers = [
		"Файл",
		"Контрагенты",
		"№ согл.",
		"Год",
		*get_localized_months_list(),  # Генерация месяцев
		"Итого",
		"Отклонение (-)",
		"Ошибки"
	]

	# Записываем заголовки в четвертую строку
	for col_num, header in enumerate(headers, 1):
		col_letter = get_column_letter(col_num)
		cell = ws[f"{col_letter}4"]
		cell.value = header
		cell.font = Font(bold=True)
		cell.alignment = Alignment(vertical='center')

		# Выравнивание для разных колонок
		if col_num == 1:  # Файл
			cell.alignment = Alignment(horizontal='left')
		elif col_num in [2, 3]:  # Контрагенты, № соглашения
			cell.alignment = Alignment(horizontal='left')
		elif col_num == 4:  # Год
			cell.alignment = Alignment(horizontal='left')
		elif 5 <= col_num <= 16:  # Месяцы 01-12
			cell.alignment = Alignment(horizontal='right')
		elif col_num == 17:  # Итого
			cell.alignment = Alignment(horizontal='right')
		elif col_num == 18:  # +/-
			cell.alignment = Alignment(horizontal='center')
		elif col_num == 19:  # Ошибки
			cell.alignment = Alignment(horizontal='left')

	# Заполняем данные (начинаем с 5 строки)
	row_num = 5

	for doc_item in documents:
		# РАЗБИРАЕМ КОРТЕЖ: (document, summary)
		if isinstance(doc_item, tuple) and len(doc_item) == 2:
			document, summary = doc_item
		else:
			# Для обратной совместимости
			document = doc_item
			summary = {}

		# Получаем имя файла безопасным способом
		file_path_obj = Path(document.file_path) if isinstance(document.file_path, str) else document.file_path
		file_name = file_path_obj.name

		customer_names = format_string_list(document.customer_names, default_text="не определен")

		base_data = {
			"file_path": str(document.file_path),
			"customer_names": customer_names,
			"agreement_number": document.agreement_number,
			"year": document.year,
			"allowed_deviation": document.allowed_deviation,
			"validation_errors": format_string_list(document.validation_errors)
		}

		# Суммируем планы по месяцам ИЗ SUMMARY (новый способ)
		monthly_totals = [0.0] * 12
		if summary:
			# Суммируем планы всех покупателей для этого документа
			for customer_plans in summary.values():
				for month_idx, value in enumerate(customer_plans):
					if value is not None:
						monthly_totals[month_idx] += value
		else:
			# Старый способ (для обратной совместимости)
			if hasattr(document, 'plans'):
				for plan in document.plans:
					if 1 <= plan.month <= 12 and plan.planned_quantity is not None:
						month_index = plan.month - 1
						monthly_totals[month_index] += plan.planned_quantity

		# Добавляем значок файла и гиперссылку (колонка A)
		cell = ws[f'A{row_num}']
		cell.value = "📄"
		cell.hyperlink = Hyperlink(
			display=f"Источник {file_name}",
			ref=f"A{row_num}",
			target=str(document.file_path),
			tooltip=f"Открыть файл: {file_name}"
		)
		cell.font = Font(color="0000FF", underline='single')
		cell.alignment = Alignment(horizontal='left', vertical='center')

		# Данные документа
		ws[f'B{row_num}'] = base_data["customer_names"]
		ws[f'B{row_num}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

		ws[f'C{row_num}'] = base_data["agreement_number"]
		ws[f'C{row_num}'].alignment = Alignment(horizontal='left')

		ws[f'D{row_num}'] = base_data["year"]
		ws[f'D{row_num}'].alignment = Alignment(horizontal='left')

		# Данные по месяцам (колонки E-P)
		total_quantity = 0
		for month_idx, month_value in enumerate(monthly_totals):
			month_col = get_column_letter(5 + month_idx)  # Колонки E-P (5-16)
			value = month_value if month_value != 0 else None
			cell = ws[f'{month_col}{row_num}']
			cell.value = value
			cell.alignment = Alignment(horizontal='center')
			total_quantity += month_value

		# Итоговая сумма (колонка Q)
		total_cell = ws[f'Q{row_num}']
		total_cell.value = total_quantity if total_quantity != 0 else None
		total_cell.alignment = Alignment(horizontal='right')
		total_cell.font = Font(bold=True)

		# Остальные данные (колонки R-S)
		ws[f'R{row_num}'] = base_data["allowed_deviation"]
		ws[f'R{row_num}'].alignment = Alignment(horizontal='center')

		ws[f'S{row_num}'] = base_data["validation_errors"]
		ws[f'S{row_num}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

		# Добавляем границы для данных
		for col_num in range(1, 20):  # Колонки A-S
			col_letter = get_column_letter(col_num)
			cell = ws[f'{col_letter}{row_num}']
			cell.border = Border(
				left=Side(style='thin'),
				right=Side(style='thin'),
				bottom=Side(style='thin')
			)

		# Если есть ошибки, красим всю строку в красный
		if document.validation_errors:
			red_font = Font(color="FF0000")
			for col_letter in ['B', 'C', 'D', 'Q', 'R', 'S']:
				ws[f'{col_letter}{row_num}'].font = red_font
			for month_col in range(5, 17):
				ws[f'{get_column_letter(month_col)}{row_num}'].font = red_font

		row_num += 1

	# Автоподбор ширины колонок (без изменений)
	column_widths = {}
	for column in ws.columns:
		max_length = 0
		column_letter = get_column_letter(column[0].column)
		for cell in column:
			try:
				if cell.value and len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
		column_widths[column_letter] = min(max_length + 2, 15)

	# Особые настройки ширины
	column_widths['A'] = 6
	column_widths['B'] = 45
	column_widths['C'] = 10
	column_widths['D'] = 6
	column_widths['Q'] = 10
	column_widths['R'] = 6
	column_widths['S'] = 20

	for col_letter, width in column_widths.items():
		ws.column_dimensions[col_letter].width = width

	# Добавляем границы для шапки
	for col_num in range(1, 20):
		col_letter = get_column_letter(col_num)
		cell = ws[f'{col_letter}4']
		cell.border = Border(
			left=Side(style='medium'),
			right=Side(style='medium'),
			top=Side(style='medium'),
			bottom=Side(style='medium')
		)

	# Создаем директорию для экспорта, если её нет
	export_dir = export_dir or Path("export")
	export_dir.mkdir(parents=True, exist_ok=True)

	# Генерируем имя файла и проверяем существующие part-файлы
	base_filename = f"export_plans_{year}"
	if force_overwrite:
		export_file_path, part_files = get_export_file_path(export_dir, base_filename, postfix)

		# Очищаем существующие файлы, если нужно
		cleanup_existing_files(export_file_path, part_files, force_overwrite)

	else:
		# Получаем уникальное имя файла с учетом force_overwrite
		export_file_path = get_unique_filename(
			export_dir,
			base_filename,
			postfix,
			".xlsx"
		)

	# Сохраняем файл
	wb.save(export_file_path)

	return export_file_path
